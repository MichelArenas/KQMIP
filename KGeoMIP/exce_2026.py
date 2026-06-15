#!/usr/bin/env python3
# exec_2026.py — Colocar en la raíz de GeoMIP

import argparse
import multiprocessing
import shutil
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURACIÓN DE RUTAS (AJUSTA SEGÚN TU ESTRUCTURA)
# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent

# Raíz de GeoMIP (donde está este script)
GEOMIP_ROOT = SCRIPT_DIR

# Raíz del proyecto QNodes (hermano de GeoMIP)
QNODES_ROOT = GEOMIP_ROOT.parent / "QNodes"

# Ruta al módulo Method2 dentro de GeoMIP
METHOD2_SRC = GEOMIP_ROOT / "src" / "Method2_Dynamic_Programming_Reformulation" / "src"

# Añadir rutas al path de Python
sys.path.insert(0, str(METHOD2_SRC))               # para imports de GeoMIP
sys.path.insert(0, str(QNODES_ROOT / "src"))       # para imports de QNodes (q_nodes_k)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTES (1‑indexed)
# ─────────────────────────────────────────────────────────────────────────────
FILA_PRIMER_DATO = 6
COL_ALCANCE = 2
COL_MECANISMO = 3

COL_INICIO_QNODES = {2: 4, 3: 10, 4: 16, 5: 22}
COL_INICIO_GEOMETRIC = {2: 7, 3: 13, 4: 19, 5: 25}

HOJAS_EXCLUIDAS = {"plataformas", "Requerimientos"}
ABECEDARY = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

# ─────────────────────────────────────────────────────────────────────────────
#  FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────────────────
def letras_a_binario(texto: str, n_bits: int) -> str:
    resultado = ["0"] * n_bits
    for letra in texto.upper():
        idx = ABECEDARY.find(letra)
        if 0 <= idx < n_bits:
            resultado[idx] = "1"
    return "".join(resultado)

def resolver_tpm(n_bits: int) -> Path:
    nombre = f"N{n_bits}A.csv"
    candidatos = [
        GEOMIP_ROOT / "data" / "samples" / nombre,
        METHOD2_SRC / ".." / ".samples" / nombre,
        QNODES_ROOT / "src" / ".samples" / nombre,
    ]
    for c in candidatos:
        if c.exists():
            return c
    raise FileNotFoundError(f"No se encontró '{nombre}'")

# ─────────────────────────────────────────────────────────────────────────────
#  WORKERS
# ─────────────────────────────────────────────────────────────────────────────
def worker_geometric(estado_ini, condicion, alcance, mecanismo, tpm, k, queue):
    try:
        if k == 2:
            from src.controllers.strategies.geometric import GeometricSIA
            from src.controllers.manager import Manager
            gestor = Manager(estado_inicial=estado_ini)
            estrategia = GeometricSIA(gestor)
            resultado = estrategia.aplicar_estrategia(condicion, alcance, mecanismo, tpm)
        else:
            from src.controllers.strategies.geometric_k import GeometricSIAK
            from src.controllers.manager import Manager
            gestor = Manager(estado_inicial=estado_ini)
            estrategia = GeometricSIAK(gestor)
            resultado = estrategia.aplicar_estrategia(condicion, alcance, mecanismo, tpm, k=k)

        queue.put({
            "particion": str(resultado.particion),
            "perdida": str(resultado.perdida).replace(".", ","),
            "tiempo": str(resultado.tiempo_ejecucion).replace(".", ","),
        })
    except Exception as e:
        import traceback
        queue.put({"particion": f"ERROR: {e}", "perdida": None, "tiempo": None, "traceback": traceback.format_exc()})

def worker_qnodes(estado_ini, condicion, alcance, mecanismo, tpm, k, queue):
    try:
        if k == 2:
            # Usar QNodes de GeoMIP (recibe gestor)
            from src.controllers.manager import Manager
            from src.controllers.strategies.q_nodes import QNodes
            gestor = Manager(estado_inicial=estado_ini)
            estrategia = QNodes(gestor)
            resultado = estrategia.aplicar_estrategia(condicion, alcance, mecanismo)
        else:
            # Usar KQNodes de QNodes (recibe tpm)
            # Asegurar que se pueda importar desde QNODES_ROOT
            sys.path.insert(0, str(QNODES_ROOT / "src"))
            from src.strategies.q_nodes_k import KQNodes
            estrategia = KQNodes(tpm)
            resultado = estrategia.aplicar_estrategia(estado_ini, condicion, alcance, mecanismo, k=k)

        queue.put({
            "particion": str(resultado.particion),
            "perdida": str(resultado.perdida).replace(".", ","),
            "tiempo": str(resultado.tiempo_ejecucion).replace(".", ","),
        })
    except Exception as e:
        import traceback
        queue.put({"particion": f"ERROR: {e}", "perdida": None, "tiempo": None, "traceback": traceback.format_exc()})

def ejecutar_con_timeout(target, args, timeout=3600):
    queue = multiprocessing.Queue()
    proc = multiprocessing.Process(target=target, args=args + (queue,))
    proc.start()
    proc.join(timeout=timeout)
    if proc.is_alive():
        proc.terminate()
        proc.join()
        return {"particion": "TIMEOUT", "perdida": None, "tiempo": None}
    if not queue.empty():
        res = queue.get()
        if "traceback" in res:
            print(f"\n    [TRACEBACK]\n{res['traceback']}")
        return res
    return {"particion": None, "perdida": None, "tiempo": None}

# ─────────────────────────────────────────────────────────────────────────────
#  PROCESAMIENTO DE HOJA
# ─────────────────────────────────────────────────────────────────────────────
def procesar_hoja(wb, sheet_name, inicio=0, cantidad=50, ks=None,
                  timeout=3600, solo_geometric=False, solo_qnodes=False):
    if ks is None:
        ks = [2, 3, 4, 5]
    ws = wb[sheet_name]

    raw_estado = ws.cell(row=2, column=2).value
    raw_sistema = ws.cell(row=3, column=2).value
    if not raw_sistema:
        print(f"  [SKIP] {sheet_name}: no hay sistema en B3.")
        return
    n_bits = len(str(raw_sistema))
    if raw_estado and len(str(raw_estado)) == n_bits:
        estado_ini = str(raw_estado)
    else:
        estado_ini = "1" + "0" * (n_bits - 1)
        print(f"  Estado inicial inferido: {estado_ini}")

    condicion = "1" * n_bits

    try:
        tpm_path = resolver_tpm(n_bits)
        tpm = np.genfromtxt(tpm_path, delimiter=",")
    except FileNotFoundError as e:
        print(f"  [ERROR TPM] {e}")
        return

    print(f"\n{'='*70}")
    print(f"Hoja: {sheet_name} | n={n_bits} | TPM: {tpm_path.name}")
    print(f"Filas {inicio+1}..{inicio+cantidad} | k={ks} | timeout={timeout}s")
    print(f"{'='*70}")

    for idx in range(inicio, inicio + cantidad):
        row = FILA_PRIMER_DATO + idx
        alcance_raw = ws.cell(row=row, column=COL_ALCANCE).value
        mecanismo_raw = ws.cell(row=row, column=COL_MECANISMO).value
        if not alcance_raw or not mecanismo_raw:
            print(f"  Fila {idx+1}: sin datos, fin de hoja.")
            break

        alcance = letras_a_binario(str(alcance_raw), n_bits)
        mecanismo = letras_a_binario(str(mecanismo_raw), n_bits)
        print(f"\n  [{idx+1}] alcance={alcance_raw}  mecanismo={mecanismo_raw}")

        for k in ks:
            if not solo_geometric:
                print(f"    QNodes  k={k} ... ", end="", flush=True)
                t0 = time.perf_counter()
                res = ejecutar_con_timeout(worker_qnodes, (estado_ini, condicion, alcance, mecanismo, tpm, k), timeout)
                dt = time.perf_counter() - t0
                print(f"φ={res['perdida']}  ({dt:.1f}s)")
                col = COL_INICIO_QNODES[k]
                ws.cell(row=row, column=col).value = res["particion"]
                ws.cell(row=row, column=col+1).value = res["perdida"]
                ws.cell(row=row, column=col+2).value = res["tiempo"]

            if not solo_qnodes:
                print(f"    Geometric k={k} ... ", end="", flush=True)
                t0 = time.perf_counter()
                res = ejecutar_con_timeout(worker_geometric, (estado_ini, condicion, alcance, mecanismo, tpm, k), timeout)
                dt = time.perf_counter() - t0
                print(f"φ={res['perdida']}  ({dt:.1f}s)")
                col = COL_INICIO_GEOMETRIC[k]
                ws.cell(row=row, column=col).value = res["particion"]
                ws.cell(row=row, column=col+1).value = res["perdida"]
                ws.cell(row=row, column=col+2).value = res["tiempo"]

    print(f"\n  Hoja '{sheet_name}' procesada.")

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--entrada", default=str(GEOMIP_ROOT / "results" / "DatosPruebas2026_1.xlsx"))
    parser.add_argument("--salida", default=str(GEOMIP_ROOT / "results" / "DatosPruebas2026_1_resultados.xlsx"))
    parser.add_argument("--hoja")
    parser.add_argument("--inicio", type=int, default=0)
    parser.add_argument("--cantidad", type=int, default=50)
    parser.add_argument("--timeout", type=int, default=3600)
    parser.add_argument("--ks", default="2,3,4,5")
    parser.add_argument("--solo-geometric", action="store_true")
    parser.add_argument("--solo-qnodes", action="store_true")
    args = parser.parse_args()

    ks = [int(k) for k in args.ks.split(",") if 2 <= int(k) <= 5]

    entrada = Path(args.entrada)
    if not entrada.exists():
        print(f"ERROR: {entrada} no existe")
        sys.exit(1)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(entrada, salida)
    print(f"Copia de trabajo: {salida}")

    wb = openpyxl.load_workbook(salida)
    hojas = [args.hoja] if args.hoja else [sh for sh in wb.sheetnames if sh not in HOJAS_EXCLUIDAS]

    for sh in hojas:
        procesar_hoja(wb, sh, inicio=args.inicio, cantidad=args.cantidad, ks=ks,
                      timeout=args.timeout, solo_geometric=args.solo_geometric, solo_qnodes=args.solo_qnodes)
        wb.save(salida)

    print(f"\n✅ Resultados guardados en {salida}")

if __name__ == "__main__":
    main()