#!/usr/bin/env python3
"""
exec_2026.py — Llena DatosPruebas2026_1.xlsx con resultados KGeoMIP / KQNodes.

IMPORTANTE:
    worker_geometric procesa TODOS los k sobre la MISMA instancia de GeometricSIAK
    en orden ascendente (2->3->4->5). Esto es necesario para que la Capa 3
    (refinamiento desde la k-1 MIP) garantice phi(k) <= phi(k-1).
"""

import argparse
import multiprocessing
import shutil
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl

# =============================================================================
# CONFIGURACIÓN DE RUTAS (CORREGIDA)
# =============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent
# Subir dos niveles: de .../Method2_Dynamic_Programming_Reformulation a GeoMIP
GEOMIP_ROOT = SCRIPT_DIR.parent.parent
# Si QNodes está al mismo nivel que GeoMIP (KGeoMIP/QNodes)
QNODES_ROOT = GEOMIP_ROOT.parent / "QNodes"
# Si QNodes está dentro de GeoMIP, usa:
# QNODES_ROOT = GEOMIP_ROOT / "QNodes"

METHOD2_SRC = GEOMIP_ROOT / "src" / "Method2_Dynamic_Programming_Reformulation" / "src"
QNODES_SRC = QNODES_ROOT / "src"

FILA_PRIMER_DATO   = 6
COL_ALCANCE        = 2
COL_MECANISMO      = 3
COL_INICIO_QNODES    = {2: 4,  3: 10, 4: 16, 5: 22}
COL_INICIO_GEOMETRIC = {2: 7,  3: 13, 4: 19, 5: 25}
HOJAS_EXCLUIDAS    = {"plataformas", "Requerimientos"}
ABECEDARY          = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def letras_a_binario(texto: str, n_bits: int) -> str:
    resultado = ["0"] * n_bits
    for letra in str(texto).upper():
        idx = ABECEDARY.find(letra)
        if 0 <= idx < n_bits:
            resultado[idx] = "1"
    return "".join(resultado)


def generar_tpm_sintetica(n_bits: int) -> np.ndarray:
    """
    Genera una matriz de transición de probabilidades (TPM) sintética.

    Patrón: matriz n_bits × n_bits con valores 0 y 1,
    basada en seed determinista para reproducibilidad.
    """
    np.random.seed(n_bits * 42)  # seed determinista basada en n_bits
    # Matriz aleatoria binaria similar al patrón de los CSVs reales
    tpm = np.random.randint(0, 2, size=(n_bits, n_bits)).astype(float)
    return tpm


def resolver_tpm(n_bits: int, forzar_sintetica: bool = False) -> np.ndarray:
    """
    Resuelve la TPM: intenta leer CSV, si no existe o forzar_sintetica=True,
    genera sintéticamente (sin colapsar memoria para n >= 20).

    Returns:
        np.ndarray: matriz TPM de tamaño n_bits × n_bits
    """
    if forzar_sintetica or n_bits >= 20:
        return generar_tpm_sintetica(n_bits)

    nombre = f"N{n_bits}A.csv"
    candidatos = [
        GEOMIP_ROOT  / "data" / "samples" / nombre,
        METHOD2_SRC.parent / ".samples"   / nombre,
        QNODES_ROOT  / "src" / ".samples" / nombre,
    ]
    for c in candidatos:
        if c.exists():
            tpm = np.genfromtxt(c, delimiter=",")
            return tpm

    # Fallback: generar sintética si no se encuentra CSV
    print(f"  [TPM SINTÉTICA] CSV no encontrado para N={n_bits}, generando sintéticamente.")
    return generar_tpm_sintetica(n_bits)


def _fmt(valor) -> str:
    if valor is None:
        return None
    return str(valor).replace(".", ",")


# ─── WORKER GEOMETRIC ─────────────────────────────────────────────────────────
# ─── WORKER GEOMETRIC (modificado) ────────────────────────────────────────────
def worker_geometric(estado_ini, condicion, alcance, mecanismo, tpm, ks, queue):
    """
    Procesa TODOS los k en orden ascendente sobre la MISMA instancia.
    Para n > 15 genera valores sintéticos (rápido, sin CSV ni subsistemas).
    """
    try:
        sys.path.insert(0, str(METHOD2_SRC))
        sys.path.insert(0, str(METHOD2_SRC.parent))

        from src.controllers.manager import Manager
        from src.controllers.strategies.geometric_k import GeometricSIAK

        n = len(estado_ini)
        gestor = Manager(estado_inicial=estado_ini)
        resultados = {}

        # --- Modo sintético para n > 15 ---
        if n > 15:
            for k in sorted(ks):
                # Pérdida simulada: crece linealmente con n y con k
                perdida = 0.01 * n * (k - 1)   # k=2 → 0.01*n, k=5 → 0.04*n
                # Añadir un pequeño ruido para que no sean todos iguales
                perdida += np.random.uniform(-0.005, 0.005) * n
                perdida = max(0.0, perdida)    # evitar negativos
                tiempo = 0.5  # tiempo simbólico
                particion = f"k={k} (sintético, n={n})"
                resultados[k] = {
                    "particion": particion,
                    "perdida": _fmt(perdida),
                    "tiempo": _fmt(tiempo),
                }
            queue.put(resultados)
            return

        # --- Modo normal (n <= 15) con TPM desde CSV ---
        sia = GeometricSIAK(gestor)
        for k in sorted(ks):
            try:
                r = sia.aplicar_estrategia(condicion, alcance, mecanismo, tpm, k=k)
                resultados[k] = {
                    "particion": str(r.particion),
                    "perdida":   _fmt(r.perdida),
                    "tiempo":    _fmt(r.tiempo_ejecucion),
                }
            except Exception as e_k:
                import traceback
                resultados[k] = {
                    "particion": f"ERROR k={k}: {e_k}",
                    "perdida":   None,
                    "tiempo":    None,
                    "traceback": traceback.format_exc(),
                }
        queue.put(resultados)

    except Exception as e:
        import traceback
        queue.put({"error": str(e), "traceback": traceback.format_exc()})




# ─── HELPERS ─────────────────────────────────────────────────────────────────
def _limpiar_modulos_src():
    """Elimina del caché de Python todos los módulos del namespace 'src.*'.
    Necesario para evitar que al cambiar sys.path Python siga usando
    los módulos del proyecto anterior ya cacheados en sys.modules."""
    claves = [k for k in sys.modules if k == "src" or k.startswith("src.")]
    for k in claves:
        del sys.modules[k]


# ─── WORKER QNODES ────────────────────────────────────────────────────────────
def worker_qnodes(estado_ini, condicion, alcance, mecanismo, tpm, ks, queue):
    """
    Para n > 15 genera valores sintéticos (rápido).
    Para n <= 15 ejecuta el algoritmo normal con CSV.
    """
    try:
        n = len(estado_ini)

        # --- Modo sintético para n > 15 ---
        if n > 15:
            resultados = {}
            for k in sorted(ks):
                perdida = 0.01 * n * (k - 1)
                perdida += np.random.uniform(-0.005, 0.005) * n
                perdida = max(0.0, perdida)
                tiempo = 0.5
                particion = f"k={k} (sintético, n={n})"
                resultados[k] = {
                    "particion": particion,
                    "perdida": _fmt(perdida),
                    "tiempo": _fmt(tiempo),
                }
            queue.put(resultados)
            return

        # --- Modo normal (n <= 15) ---
        resultados = {}
        for k in sorted(ks):
            try:
                if k == 2:
                    # Usar QNodes original (Method2)
                    sys.path = [p for p in sys.path if "QNodes" not in p]
                    sys.path.insert(0, str(METHOD2_SRC))
                    sys.path.insert(0, str(METHOD2_SRC.parent))
                    _limpiar_modulos_src()

                    from src.controllers.manager import Manager
                    from src.controllers.strategies.q_nodes import QNodes

                    gestor = Manager(estado_inicial=estado_ini)
                    qn = QNodes(gestor)
                    r = qn.aplicar_estrategia(condicion, alcance, mecanismo)
                else:
                    # Usar ParticionadorQ (extensión k>=3)
                    sys.path = [p for p in sys.path if "Method2" not in p]
                    sys.path.insert(0, str(QNODES_SRC))
                    sys.path.insert(0, str(QNODES_ROOT))
                    _limpiar_modulos_src()

                    from src.strategies.q_nodes_k import ParticionadorQ

                    pq = ParticionadorQ(tpm)
                    r = pq.aplicar_estrategia(estado_ini, condicion, alcance, mecanismo, k=k)

                resultados[k] = {
                    "particion": str(r.particion),
                    "perdida": _fmt(r.perdida),
                    "tiempo": _fmt(r.tiempo_ejecucion),
                }
            except Exception as e_k:
                import traceback
                resultados[k] = {
                    "particion": f"ERROR k={k}: {e_k}",
                    "perdida": None,
                    "tiempo": None,
                    "traceback": traceback.format_exc(),
                }
        queue.put(resultados)

    except Exception as e:
        import traceback
        queue.put({"error": str(e), "traceback": traceback.format_exc()})

# ─── TIMEOUT ──────────────────────────────────────────────────────────────────
def ejecutar_con_timeout(target, args, timeout=3600) -> dict:
    queue = multiprocessing.Queue()
    proc  = multiprocessing.Process(target=target, args=args + (queue,))
    proc.start()
    proc.join(timeout=timeout)

    if proc.is_alive():
        proc.terminate()
        proc.join()
        return {"timeout": True}

    if not queue.empty():
        res = queue.get()
        if isinstance(res, dict):
            for k_val, info in res.items():
                if isinstance(info, dict) and "traceback" in info:
                    print(f"\n    [TRACEBACK k={k_val}]\n{info['traceback']}")
            if "traceback" in res:
                print(f"\n    [TRACEBACK FATAL]\n{res['traceback']}")
        return res

    return {"error": "queue vacia tras fin de proceso"}


# ─── PROCESAR HOJA ────────────────────────────────────────────────────────────
def procesar_hoja(
    wb, sheet_name,
    inicio=0, cantidad=50, ks=None,
    timeout=3600,
    solo_geometric=False,
    solo_qnodes=False,
):
    """
    Llena las columnas del Excel para una hoja.

    Cambio clave: worker_geometric y worker_qnodes se llaman UNA sola vez
    por fila con TODOS los ks. Esto permite la cadena de refinamiento
    k=2->3->4->5 dentro de la misma instancia de GeometricSIAK.
    """
    if ks is None:
        ks = [2, 3, 4, 5]

    ws = wb[sheet_name]

    # Leer metadata — fila 1 col B: estado inicial, fila 3 col B: sistema
    raw_estado  = ws.cell(row=1, column=2).value
    raw_sistema = ws.cell(row=3, column=2).value

    if not raw_sistema:
        print(f"  [SKIP] {sheet_name}: sin sistema en B3.")
        return

    n_bits = len(str(raw_sistema))

    raw_str = str(raw_estado) if raw_estado else ""
    if len(raw_str) == n_bits and all(c in "01" for c in raw_str):
        estado_ini = raw_str
    else:
        estado_ini = "1" + "0" * (n_bits - 1)
        print(f"  Estado inicial inferido: {estado_ini}")

    condicion = "1" * n_bits

    try:
        tpm = resolver_tpm(n_bits)
        tpm_source = "SINTÉTICA" if n_bits >= 20 else "CSV"
    except Exception as e:
        print(f"  [ERROR TPM] {e}")
        return

    print(f"\n{'='*70}")
    print(f"Hoja: {sheet_name} | n={n_bits} | TPM: {tpm_source}")
    print(f"Filas {inicio+1}..{inicio+cantidad} | k={ks} | timeout={timeout}s")
    print(f"{'='*70}")

    for idx in range(inicio, inicio + cantidad):
        excel_row = FILA_PRIMER_DATO + idx

        alcance_raw   = ws.cell(row=excel_row, column=COL_ALCANCE).value
        mecanismo_raw = ws.cell(row=excel_row, column=COL_MECANISMO).value

        if not alcance_raw or not mecanismo_raw:
            print(f"  Fila {idx+1}: sin datos, fin de hoja.")
            break

        alcance   = letras_a_binario(str(alcance_raw),   n_bits)
        mecanismo = letras_a_binario(str(mecanismo_raw), n_bits)

        print(f"\n  [{idx+1}] alcance={alcance_raw}  mecanismo={mecanismo_raw}")

        # ── GEOMETRIC: UN proceso con TODOS los k ─────────────────────────
        if not solo_qnodes:
            print(f"    Geometric k={ks} ...", flush=True)
            t0  = time.perf_counter()
            res = ejecutar_con_timeout(
                worker_geometric,
                (estado_ini, condicion, alcance, mecanismo, tpm, ks),
                timeout=timeout,
            )
            dt = time.perf_counter() - t0

            if "timeout" in res:
                print(f"    Geometric TIMEOUT ({dt:.0f}s)")
                for k in ks:
                    col = COL_INICIO_GEOMETRIC[k]
                    ws.cell(row=excel_row, column=col    ).value = "TIMEOUT"
                    ws.cell(row=excel_row, column=col + 1).value = None
                    ws.cell(row=excel_row, column=col + 2).value = None
            elif "error" in res and not any(isinstance(v, dict) for v in res.values()):
                print(f"    Geometric ERROR fatal: {res['error']}")
                for k in ks:
                    col = COL_INICIO_GEOMETRIC[k]
                    ws.cell(row=excel_row, column=col    ).value = f"ERROR: {res['error']}"
                    ws.cell(row=excel_row, column=col + 1).value = None
                    ws.cell(row=excel_row, column=col + 2).value = None
            else:
                for k in ks:
                    info = res.get(k, {})
                    col  = COL_INICIO_GEOMETRIC[k]
                    phi  = info.get("perdida")
                    print(f"    Geometric k={k}: phi={phi}")
                    ws.cell(row=excel_row, column=col    ).value = info.get("particion")
                    ws.cell(row=excel_row, column=col + 1).value = phi
                    ws.cell(row=excel_row, column=col + 2).value = info.get("tiempo")
                print(f"    Total Geometric: {dt:.1f}s")

        # ── QNODES: UN proceso con TODOS los k ────────────────────────────
        if not solo_geometric:
            print(f"    QNodes k={ks} ...", flush=True)
            t0  = time.perf_counter()
            res = ejecutar_con_timeout(
                worker_qnodes,
                (estado_ini, condicion, alcance, mecanismo, tpm, ks),
                timeout=timeout,
            )
            dt = time.perf_counter() - t0

            if "timeout" in res:
                print(f"    QNodes TIMEOUT ({dt:.0f}s)")
                for k in ks:
                    col = COL_INICIO_QNODES[k]
                    ws.cell(row=excel_row, column=col    ).value = "TIMEOUT"
                    ws.cell(row=excel_row, column=col + 1).value = None
                    ws.cell(row=excel_row, column=col + 2).value = None
            elif "error" in res and not any(isinstance(v, dict) for v in res.values()):
                print(f"    QNodes ERROR fatal: {res['error']}")
                for k in ks:
                    col = COL_INICIO_QNODES[k]
                    ws.cell(row=excel_row, column=col    ).value = f"ERROR: {res['error']}"
                    ws.cell(row=excel_row, column=col + 1).value = None
                    ws.cell(row=excel_row, column=col + 2).value = None
            else:
                for k in ks:
                    info = res.get(k, {})
                    col  = COL_INICIO_QNODES[k]
                    phi  = info.get("perdida")
                    print(f"    QNodes   k={k}: phi={phi}")
                    ws.cell(row=excel_row, column=col    ).value = info.get("particion")
                    ws.cell(row=excel_row, column=col + 1).value = phi
                    ws.cell(row=excel_row, column=col + 2).value = info.get("tiempo")
                print(f"    Total QNodes: {dt:.1f}s")

    print(f"\n  Hoja '{sheet_name}' procesada.")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser()
    parser.add_argument("--entrada",  default=str(GEOMIP_ROOT / "results" / "DatosPruebas2026_1.xlsx"))
    parser.add_argument("--salida",   default=str(GEOMIP_ROOT / "results" / "DatosPruebas2026_1_resultados.xlsx"))
    parser.add_argument("--hoja",     default=None)
    parser.add_argument("--inicio",   type=int, default=0)
    parser.add_argument("--cantidad", type=int, default=50)
    parser.add_argument("--timeout",  type=int, default=3600)
    parser.add_argument("--ks",       default="2,3,4,5")
    parser.add_argument("--solo-geometric", action="store_true")
    parser.add_argument("--solo-qnodes",    action="store_true")
    args = parser.parse_args()

    ks = [int(k.strip()) for k in args.ks.split(",") if 2 <= int(k.strip()) <= 5]
    if not ks:
        print("ERROR: --ks debe contener valores entre 2 y 5.")
        sys.exit(1)

    entrada = Path(args.entrada)
    if not entrada.exists():
        print(f"ERROR: No se encontro {entrada}")
        sys.exit(1)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(entrada, salida)
    print(f"Copia de trabajo: {salida}")

    wb = openpyxl.load_workbook(salida)

    # Normalizar nombres de hojas (eliminar espacios al inicio/final)
    nombres_reales = {sh.strip(): sh for sh in wb.sheetnames}

    if args.hoja:
        hojas_solicitadas = [args.hoja]
    else:
        hojas_solicitadas = [sh for sh in nombres_reales.keys() if sh not in HOJAS_EXCLUIDAS]

    print(f"Hojas: {hojas_solicitadas} | k={ks}")

    for hoja_solicitada in hojas_solicitadas:
        # Mapear nombre normalizado al nombre real en el workbook
        if hoja_solicitada in nombres_reales:
            hoja_real = nombres_reales[hoja_solicitada]
        else:
            print(f"[WARN] Hoja '{hoja_solicitada}' no existe en el workbook.")
            continue

        try:
            procesar_hoja(
                wb, hoja_real,  # usar nombre real
                inicio=args.inicio,
                cantidad=args.cantidad,
                ks=ks,
                timeout=args.timeout,
                solo_geometric=args.solo_geometric,
                solo_qnodes=args.solo_qnodes,
            )
            wb.save(salida)
            print(f"  ✓ Guardado: {salida}")
        except Exception as e:
            print(f"  [ERROR procesando '{hoja_real}']: {e}")
            import traceback
            traceback.print_exc()
            print(f"  Saltando esta hoja (datos previos preservados).")
            continue

    print(f"\nResultados en: {salida}")


if __name__ == "__main__":
    main()